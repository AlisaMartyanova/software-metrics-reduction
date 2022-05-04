import pandas as pd
import numpy as np
from openpyxl import load_workbook
from sklearn import preprocessing, metrics
from sklearn.decomposition import NMF


# normalize the entire dataset using min-max scaler
def normalize(dataset):
    min_max_scaler = preprocessing.MinMaxScaler()
    x_scaled = min_max_scaler.fit_transform(dataset)

    return x_scaled


# Estimate performance of the model on the data
def get_score(model, data, scorer=metrics.explained_variance_score):
    prediction = model.inverse_transform(model.transform(data))
    return scorer(data, prediction)


# singular value decomposition
def nmf(data, comps, initial_feature_names):
    model = NMF(n_components=comps, init='nndsvd', max_iter=500).fit(data)

    # get the index of the most important feature on EACH component
    most_important = [np.abs(model.components_[i]).argmax() for i in range(comps)]

    # get the names
    most_important_names = [initial_feature_names[most_important[i]] for i in range(comps)]
    return most_important_names, get_score(model, data)*100


# save results as a table
def interpret_results(initial_names, xlsx_filename, result_names, variance):
    wb = load_workbook(xlsx_filename)
    sheet = wb.active
    sheet.cell(1, 1, 'Number of metrics')
    sheet.cell(1, 2, 'Variance explained')

    for i in range(len(initial_names)):
        sheet.cell(1, i + 3, initial_names[i])
    for i in range(len(result_names)):
        sheet.cell(i+2, 1, i+1)
        sheet.cell(i+2, 2, variance[i])
        for r in result_names[i]:
            for c in range(len(initial_names)):
                if initial_names[c] == r:
                    if not sheet.cell(i+2, c+3).value:
                        sheet.cell(i + 2, c + 3, 0)
                    sheet.cell(i+2, c+3, int(sheet.cell(i+2, c+3).value)+1)

    wb.save(xlsx_filename)


def main():

    # data with 49 metrics
    data_filename = '49.csv'

    # read the data, fill missing values with median and convert all to float
    data = pd.read_csv(data_filename)
    for i in data:
        data[i] = data[i].astype(float)
    initial_feature_names = [d for d in data]
    data = normalize(data)
    names = []
    variance = []
    for i in range(1, len(initial_feature_names)+1):
        name, var = nmf(data, i, initial_feature_names)
        names.append(name)
        variance.append(var)

    interpret_results(initial_feature_names, 'results/NMF_results_49.xlsx', names, variance)


main()
