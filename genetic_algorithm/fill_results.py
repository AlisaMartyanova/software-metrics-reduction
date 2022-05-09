from openpyxl import Workbook, load_workbook
import pandas as pd
import gen_algo

sheet_name = 'results.xlsx'
wb = load_workbook(sheet_name)
sheet = wb.active

DATA = pd.read_csv('49.csv')
columns = list(DATA.columns)

for i in range(1):
    names = gen_algo.start(i)

    sheet.cell(i, 1, )