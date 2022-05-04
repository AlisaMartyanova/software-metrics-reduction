import pandas as pd
import numpy as np
from openpyxl import load_workbook
from sklearn import preprocessing
from sklearn.decomposition import TruncatedSVD


# normalize the entire dataset using min-max scaler
def normalize(dataset):
    min_max_scaler = preprocessing.MinMaxScaler()
    # min_max_scaler = preprocessing.StandardScaler()
    x_scaled = min_max_scaler.fit_transform(dataset)

    return x_scaled


# singular value decomposition
def svd(data, comps, initial_feature_names):
    model = TruncatedSVD(n_components=comps).fit(data)
    model.transform(data)
    # print(model.explained_variance_ratio_)

    # get variance for all components
    variance = [sum(model.explained_variance_ratio_[:i]*100) for i in range(1, len(model.explained_variance_ratio_) + 1)]

    # get the index of the most important feature on EACH component
    # LIST COMPREHENSION HERE
    most_important = [np.abs(model.components_[i]).argmax() for i in range(len(model.explained_variance_ratio_))]

    # get the names
    most_important_names = [initial_feature_names[most_important[i]] for i in range(len(model.explained_variance_ratio_))]
    return most_important_names, variance


# save results as a table
def interpret_results(initial_names, xlsx_filename, result_names, variance):
    wb = load_workbook(xlsx_filename)
    sheet = wb.active
    sheet.cell(1, 1, 'Number of metrics')
    sheet.cell(1, 2, 'Variance explained')

    for i in range(len(initial_names)):
        sheet.cell(1, i + 3, initial_names[i])
    for i in range(len(result_names)):
        sheet.cell(i+2, 1, i+1)
        sheet.cell(i+2, 2, variance[i])
        for c in range(len(initial_names)):
            if initial_names[c] == result_names[i]:
                for j in range(i, len(result_names)):
                    if not sheet.cell(j+2, c+3).value:
                        sheet.cell(j + 2, c + 3, 0)
                    sheet.cell(j+2, c+3, int(sheet.cell(j+2, c+3).value)+1)

    wb.save(xlsx_filename)


def main():

    # data with 49 metrics
    data_filename = '49.csv'

    # read the data, fill missing values with median and convert all to float
    data = pd.read_csv(data_filename)
    for i in data:
        data[i] = data[i].astype(float)
    initial_feature_names = [d for d in data]
    data = normalize(data)

    components = len(initial_feature_names)

    names, variance = svd(data, components-1, initial_feature_names)

    interpret_results(initial_feature_names, 'results/SVD_results_49.xlsx', names, variance)


main()
