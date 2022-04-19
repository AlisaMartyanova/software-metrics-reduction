from openpyxl import load_workbook
from sklearn.decomposition import PCA
import pandas as pd
import numpy as np


# normalize the entire dataset using min-max scaler
def normalize(dataset):
    column_maxes = dataset.max()
    df_max = column_maxes.max()
    normalized_dataset = dataset / df_max

    return normalized_dataset


# principle component analyses
def pca(data, comps, initial_feature_names):
    model = PCA(n_components=comps).fit(data)
    model.transform(data)

    # number of components
    n_pcs = comps

    # get the index of the most important feature on EACH component
    # LIST COMPREHENSION HERE
    most_important = [np.abs(model.components_[i]).argmax() for i in range(n_pcs)]

    # get the names
    most_important_names = [initial_feature_names[most_important[i]] for i in range(n_pcs)]
    return most_important_names


# save results as a table
def interpret_results(initial_names, xlsx_filename, result_names):
    wb = load_workbook(xlsx_filename)
    sheet = wb.active
    sheet.cell(1, 1, 'Number of metrics')

    for i in range(len(result_names)):
        sheet.cell(1, i+2, initial_names[i])
        sheet.cell(i+2, 1, i+1)
        for c in range(len(initial_names)):
            if initial_names[c] == result_names[i]:
                for j in range(i, len(result_names)):
                    sheet.cell(j+2, c+2, 1)

    wb.save(xlsx_filename)


def main():

    # data with 49 metrics
    data_filename = '49.csv'

    # read the data, fill missing values with median and convert all to float
    data = pd.read_csv(data_filename)
    for i in data:
        data[i] = data[i].astype(float)
    data = normalize(data)
    initial_feature_names = [d for d in data]
    original = data.to_numpy()

    components = len(initial_feature_names)

    names = pca(original, components, initial_feature_names)
    print(len(initial_feature_names), initial_feature_names)
    print('\n')
    print(len(names), names)

    interpret_results(initial_feature_names, 'PCA_results_49.xlsx', names)


main()
