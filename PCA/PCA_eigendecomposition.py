from openpyxl import load_workbook
from sklearn import preprocessing
import pandas as pd
import numpy as np


# normalize the entire dataset using min-max scaler
def normalize(dataset):
    min_max_scaler = preprocessing.MinMaxScaler()
    x_scaled = min_max_scaler.fit_transform(dataset)

    return x_scaled


def pca(data, initial_feature_names):
    # calculate covariance matrix
    cov_mat = np.cov(data, rowvar=False)

    # from this covariance matrix, calculate the eigenvalues and the eigenvectors
    eigen_vals, eigen_vecs = np.linalg.eig(cov_mat)

    # sort eigenvalues and eigenvectors
    sorted_index = np.argsort(eigen_vals)[::-1]
    sorted_eigenvalue = eigen_vals[sorted_index]
    sorted_eigenvectors = eigen_vecs[:, sorted_index]

    # get the variance of all components
    variance_explained = []
    for i in sorted_eigenvalue:
        variance_explained.append((i / sum(sorted_eigenvalue)) * 100)
    variance = [sum(variance_explained[:i]) for i in range(1, len(variance_explained)+1)]

    # get most important feature from each component
    most_important = [np.abs(sorted_eigenvectors[i]).argmax() for i in range(len(sorted_eigenvectors))]
    most_important_names = [initial_feature_names[i] for i in most_important[:len(sorted_eigenvectors)]]
    return most_important_names, variance


# save results as a table
def interpret_results(initial_names, xlsx_filename, result_names, variance):
    wb = load_workbook(xlsx_filename)
    sheet = wb.active
    sheet.cell(1, 1, 'Number of metrics')
    sheet.cell(1, 2, 'Variance explained')

    for i in range(len(initial_names)):
        sheet.cell(1, i + 3, initial_names[i])
    for i in range(len(result_names)):
        sheet.cell(i+2, 1, i+1)
        sheet.cell(i+2, 2, variance[i])
        for c in range(len(initial_names)):
            if initial_names[c] == result_names[i]:
                for j in range(i, len(result_names)):
                    if not sheet.cell(j+2, c+3).value:
                        sheet.cell(j + 2, c + 3, 0)
                    sheet.cell(j+2, c+3, int(sheet.cell(j+2, c+3).value)+1)

    wb.save(xlsx_filename)


def main():

    # data with 49 metrics
    data_filename = '49.csv'

    # read the data, fill missing values with median and convert all to float
    data = pd.read_csv(data_filename)
    for i in data:
        data[i] = data[i].astype(float)
    initial_feature_names = [d for d in data]
    data = normalize(data)

    names, variance = pca(data, initial_feature_names)

    interpret_results(initial_feature_names, 'results/PCA_results_49.xlsx', names, variance)


main()
