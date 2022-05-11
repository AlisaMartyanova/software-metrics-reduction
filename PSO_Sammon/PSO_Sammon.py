import json
import random
import numpy as np
import pandas as pd
from operator import itemgetter
from openpyxl import load_workbook
from sklearn import preprocessing

global or_dist_sum, dist_or_matrix


class Particle:
    def __init__(self, size):
        self.position = []          # particle position
        self.velocity = []          # particle velocity
        self.pos_best = []          # best position individual
        self.err_best = -1          # best error individual
        self.err = -1               # error individual

        for i in range(size):
            self.velocity.append(random.uniform(-1, 1))
            self.position.append(random.uniform(0, 1))

    # evaluate current fitness
    def evaluate(self, func, original, select_num):
        pos = project_position(original, self.position, select_num)
        self.err = func(pos)

        # check to see if the current position is an individual best
        if self.err < self.err_best or self.err_best == -1:
            self.pos_best = self.position
            self.err_best = self.err

    # update new particle velocity
    def update_velocity(self, global_pos_best, size):
        w = 0.6       # constant inertia weight (how much to weigh the previous velocity)
        c1 = 1        # cognative constant
        c2 = 2        # social constant

        for i in range(size):
            r1 = random.random()
            r2 = random.random()

            vel_cognitive = c1*r1*(self.pos_best[i]-self.position[i])
            vel_social = c2*r2*(global_pos_best[i]-self.position[i])
            self.velocity[i] = w*self.velocity[i]+vel_cognitive+vel_social

    # update the particle position based off new velocity updates
    def update_position(self, bounds, size):
        for i in range(0, size):
            self.position[i] = self.position[i] + self.velocity[i]

            # adjust maximum position if necessary
            if self.position[i] > bounds[i][1]:
                self.position[i] = bounds[i][1]

            # adjust minimum position if neseccary
            if self.position[i] < bounds[i][0]:
                self.position[i] = bounds[i][0]


class PSO:
    def __init__(self, func, original, bounds, num_particles, maxiter, select_num):

        num_dimensions = len(original[0])
        self.err_best_g = -1  # best error for group
        self.pos_best_g = []
        self.select_num = select_num

        # establish the swarm
        swarm = []
        for i in range(num_particles):
            swarm.append(Particle(num_dimensions))

        # begin optimization loop
        i = 0
        while i < maxiter:
            # print('iteration: ', i)
            # print i,err_best_g
            # cycle through particles in swarm and evaluate fitness
            for j in range(num_particles):
                swarm[j].evaluate(func, original, select_num)

                # determine if current particle is the best (globally)
                if swarm[j].err < self.err_best_g or self.err_best_g == -1:
                    self.pos_best_g = list(swarm[j].position)
                    self.err_best_g = float(swarm[j].err)

            # cycle through swarm and update velocities and position
            for j in range(num_particles):
                swarm[j].update_velocity(self.pos_best_g, num_dimensions)
                swarm[j].update_position(bounds, num_dimensions)
            i += 1

    def get_result(self):
        position = np.array(self.pos_best_g)
        sorted_index_arr = np.argsort(position)
        rslt = sorted_index_arr[-self.select_num:]
        rslt.sort()
        return {
            'features_num': self.select_num,
            'features_idxs': rslt.tolist(),
            'error': self.err_best_g
        }


def sammon_error(projected):
    global dist_or_matrix, or_dist_sum
    diff_sum = 0
    for i in range(len(projected)):
        for j in range(i + 1, len(projected)):
            diff_sum += ((dist_or_matrix[i][j] - np.linalg.norm(projected[i] - projected[j])) ** 2) / dist_or_matrix[i][j]
    return (1 / or_dist_sum) * diff_sum


def project_position(original, position, select_num):
    position = np.array(position)
    sorted_index_arr = np.argsort(position)
    rslt = sorted_index_arr[-select_num:]
    rslt.sort()
    projection = original[:, rslt]
    return projection


def interpret_results(json_filename, data_filename, xlsx_filename):
    data = pd.read_csv(data_filename)
    with open(json_filename) as json_file:
        json_results = json.load(json_file)
    wb = load_workbook(xlsx_filename)
    sheet = wb.active
    sheet.cell(1, 1, 'Number of metrics')
    sheet.cell(1, 2, 'Error')
    i = 3
    for d in data:
        sheet.cell(1, i, d)
        i += 1
    for i in range(len(json_results)):
        sheet.cell(i+2, 1, json_results[i]['features_num'])
        sheet.cell(i+2, 2, json_results[i]['error'])
        for j in json_results[i]['features_idxs']:
            sheet.cell(i + 2, j+3, 'X')
    wb.save(xlsx_filename)


# normalize the entire dataset using min-max scaler
def normalize(dataset):
    min_max_scaler = preprocessing.MinMaxScaler()
    x_scaled = min_max_scaler.fit_transform(dataset)

    return x_scaled


def main():
    data_filename = '49.csv'
    iterations = 100
    population = 50

    # read the data, fill missing values with median and convert all to float
    data = pd.read_csv(data_filename)
    # data.fillna(data.median().round(1), inplace=True)
    for i in data:
        data[i] = data[i].astype(float)
    # original = normalize(data)
    # original.to_csv('normalized_49.csv')
    original = data.to_numpy()

    # calculate distance in original matrix
    global or_dist_sum, dist_or_matrix
    or_dist_sum = 0
    dist_or_matrix = [[0 for x in range(len(original))] for y in range(len(original))]
    for i in range(len(original)):
        for j in range(i + 1, len(original)):
            dist_or = np.linalg.norm(original[i] - original[j])
            dist_or_matrix[i][j] = dist_or
            or_dist_sum += dist_or

    # bounds for feature values where 0 - non selected feature, 1 - selected
    bounds = [(0, 1) for i in range(len(original[0]))]

    # result is array with dicts with data for sets with 2-50 features
    res = []

    # check all sets with number of features from 2 to 50
    for i in range(2, len(original[0])):
        print('features num: ', i)
        pso = PSO(sammon_error, original, bounds, population, iterations, i)
        temp = pso.get_result()
        print(temp)
        print('\n')
        res.append(temp)

    res = sorted(res, key=itemgetter('error'))

    with open('results/PSO_Sammon_results_49.json', 'w') as outfile:
        json.dump(res, outfile)


main()
interpret_results('results/PSO_Sammon_results_49.json', '49.csv', 'results/PSO_Sammon_results_49.xlsx')

